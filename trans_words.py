import requests
import time
from bs4 import BeautifulSoup
from openpyxl import *


wb = load_workbook('Eng.xlsx')
sheet = wb['Лист1']


headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'}


def delete_xa0_from_excel(words: list) -> list:
	'''Удаляет xa0 из файла excel'''
	sorted_words = []
	for word in words:
		if word.value:
			if word.value.rstrip('\xa0'):
				sorted_words.append(word)

	return sorted_words


english_cells = delete_xa0_from_excel(list(sheet['A']))


def make_normal_word(word: str):
	symbols = list('йцукенгшщзхъфывапрролджэячсмитьбюё')
	word = word.lstrip(' [ ').rstrip(' ] ')

	for symbol in word:
		if not symbol in symbols:
			word = word.replace(symbol, '')

	return word


def find_description(word: str) -> list:
	'''Берет описание слова с сайта LongMan'''
	full_page = requests.get(
		'https://www.ldoceonline.com/dictionary/'+word, 
		headers=headers
	)

	soup = BeautifulSoup(full_page.content, 'html.parser')

	content = soup.findAll('span', {'class': 'DEF'})

	descriptions = []
	for index in list(range(len(content))):
		descriptions.append(content[index].text)

	return descriptions


def translate(word: str) -> list:
	'''Берет перевод(ы) слова с сайта Glosbe'''
	full_page = requests.get(
		'https://ru.glosbe.com/en/ru/'+word, headers=headers
	)

	soup = BeautifulSoup(full_page.content, 'html.parser')

	content = soup.findAll('div', {'class': 'translate-entry-translation-accents'}) 

	translated_words = []

	index = 0
	while index < 3:
		try:
			translated_words.append(make_normal_word(content[index].text))
			index += 1
		except IndexError:
			break

	return translated_words


def write_to_excel_translated_word(rus_words: list, row: int) -> None:
	'''Записывает в excel файл переведенные слова'''
	index = len(rus_words) - 1
	translate = ''

	for word in rus_words:	
		if rus_words[len(rus_words)-1] == word:
			translate += str(word)
		else:
			translate += str(word) + ', '

	sheet.cell(row=row, column=2).value = translate
		

	wb.save("Eng.xlsx")


def write_to_excel_descriptions(descriptions: list, row: int) -> None:
	'''Записывает в excel файл описание(ния) слова'''
	index = len(descriptions) - 1
	descriptions_str = ''

	for description in descriptions:

		if descriptions[len(descriptions)-1] == description:
			descriptions_str += str(description)
		else:
			description = str(
				description[1]
			).upper() + str(description[2:])
			descriptions_str += description + '. '		

	sheet.cell(row=row, column=3).value = descriptions_str

	wb.save("Eng.xlsx")


if __name__ == '__main__':
	for word in english_cells:
		if not sheet.cell(row=word.row, column=2).value:

			write_to_excel_translated_word(translate(word.value), word.row)
			write_to_excel_descriptions(find_description(word.value), word.row)

	print('Операция выполнена успешно! Иди проверь свой excel файлик)')

	time.sleep(600)