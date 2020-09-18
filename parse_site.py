import requests
import time
from bs4 import BeautifulSoup
from openpyxl import *


wb = load_workbook('Eng.xlsx')
sheet = wb['Лист1']

english_words = list(sheet['A'])
russian_words = list(sheet['C'])


ARTICLE = 'https://edition.cnn.com/2020/09/01/politics/donald-trump-race-violence-joe-biden-kenosha-election-2020/index.html'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'}


def maker_normal_words(words):
	symbols = list('йцукенгшщзхъфывапрролджэячсмитьбюё')
	checked_words = []

	for word in words:
		word = word.replace('[','').replace(']','').replace(' ','')

		for symbol in word:
			if symbol in symbols:
			    continue
			else:
			    word = word.replace(symbol,'')	

		checked_words.append(word)	

	return checked_words


def get_eng_words():
    full_page = requests.get(ARTICLE, headers=headers)

    soup = BeautifulSoup(full_page.content, 'html.parser')

    content = soup.findAll('div', {'class': 'pg-right-rail-tall', 'class': 'pg-wrapper'})
    return content[0].text


def word_from_longman(link):
    full_page = requests.get(link, headers=headers)

    soup = BeautifulSoup(full_page.content, 'html.parser')

    content = soup.findAll('h1', {'class': 'pagetitle'})
    return content[0].text


def translated_word(link):
	full_page = requests.get(link, headers=headers)

	soup = BeautifulSoup(full_page.content, 'html.parser')

	content = soup.findAll('div', {'class': 'translate-entry-translation-accents'}) 

	translated_words = []

	try:
		translated_words.append(content[2].text)
		translated_words.append(content[1].text)
		translated_words.append(content[0].text)
	except IndexError:
		try:
			translated_words.append(content[1].text)
			translated_words.append(content[0].text)
		except IndexError:
			try:
				translated_words.append(content[0].text)
			except IndexError:
				raise IndexError

	translated_words = maker_normal_words(translated_words) 

	return translated_words


words = get_eng_words().split(' ')
	

def check_this_fucking_excel(english_words):
	new_english_words = []

	for word in english_words:
		if str(type(word.value)) == "<class 'NoneType'>":
			continue
		else:
			new_english_words.append(word)

	return new_english_words



def filter_of_words(word):
    if '<' in list(word) or '>' in list(word) or '/' in list(word) or '{' in list(word) or '}' in list(word):
        return False
    elif '[' in list(word) or ']' in list(word) or '1' in list(word) or '2' in list(word) or '3' in list(word):
    	return False
    elif '4' in list(word) or '5' in list(word) or '6' in list(word) or '7' in list(word) or '8' in list(word):
    	return False
    elif '9' in list(word) or '\\' in list(word) or '-' in list(word) or '=' in list(word) or '\'' in list(word):
    	return False
    elif ';' in list(word) or '$' in list(word) or '&' in list(word) or ',' in list(word) or '@' in list(word):
    	return False
    elif '|' in list(word) or '(' in list(word) or ')' in list(word) or '0' in list(word) or '.' in list(word):
    	return False
    elif word.lower() != word or '!' in list(word) or ' ' in list(word) or '+' in list(word) or '_' in list(word):
    	return False
    elif len(word) <= 3 or '*' in list(word) or '`' in list(word) or '^' in list(word):
    	return False
    else:
        return True


def filter_for_dot(words):
	filtered_words = []
	
	for word in words:
		if word.startswith('.'):
			word = word.lstrip('.')
			filtered_words.append(word.lower())

		elif word.startswith(','):
			word = word.lstrip(',')
			iltered_words.append(word.lower())

		elif word.startswith(':'):
			word = word.lstrip(':')
			filtered_words.append(word.lower())

		elif word.startswith(';'):
			word = word.lstrip(';')
			filtered_words.append(word.lower())

		elif word.startswith('\"'):
			word = word.lstrip('\"')
			filtered_words.append(word.lower())

		elif word.startswith(' '):
			word = word.lstrip(' ')
			filtered_words.append(word.lower())

		elif word.endswith('?'):
			word = word.rstrip('?')
			filtered_words.append(word.lower())

		elif word.startswith('?'):
			word = word.lstrip('?')
			filtered_words.append(word.lower())

		elif word.endswith('.'):
			word = word.rstrip('.')
			filtered_words.append(word.lower())

		elif word.endswith(','):
			word = word.rstrip(',')
			filtered_words.append(word.lower())

		elif word.endswith(' '):
			word = word.rstrip(' ')
			filtered_words.append(word.lower())

		elif word.endswith(':'):
			word = word.rstrip(':')
			filtered_words.append(word.lower())

		elif word.endswith(';'):
			word = word.rstrip(';')
			filtered_words.append(word.lower())

		elif word.endswith('\"'):
			word = word.rstrip('\"')
			filtered_words.append(word.lower())

		else:
			filtered_words.append(word.lower())


	return filtered_words


def cheak_repeat(words):
    repeat = []

    for word in words:
        if word in repeat:
            continue
        else:
        	word = word.lower()
        	repeat.append(word)

    return repeat


def make_from_cell_list(english_words):
	new_english_words = []

	for word in english_words:
		word = word.value
		if word.endswith('\xa0'):
			word = word.rstrip('\xa0')
			new_english_words.append(word)
		else:
			new_english_words.append(word)

	return new_english_words


def cheak_repeat_in_excel(args_1, args_2):
	new_args = []
	for word in args_1:
		if word not in args_2:
			new_args.append(word)
	return new_args


def ask(words):
	new_words = []
	count_words = len(words)
	for word in words:
		inp = input('Знаешь слово: ' + word + '? ').lower()
		if inp == 'y':
			count_words -= 1
			print('Молодец! Еще ' + str(count_words) + ' слов')
			print('-----------------------------')
			continue
		else:
			print('Мы добавили это слово в поиск')
			count_words -= 1
			print('Еще ' + str(count_words) + ' слов')
			print('-----------------------------')
			new_words.append(word)

	return new_words


def words_from_longman(words):
	new_words = []
	count_words = len(words)
	for word in words:
		try:
			link = 'https://www.ldoceonline.com/search/english/direct/?q={}'.format(word)
			word = word_from_longman(link)
			new_words.append(word)
			count_words -= 1
			print('Еще ' + str(count_words) + ' слов')
			print('-----------------------------')
		except IndexError:
			count_words -= 1
			continue

	return new_words


def pk(words):
	pk = 1
	for word in words:
		pk += 1
	return pk


list_of_pk = []
def translated_words(words, pk):
	new_words = []

	for word in words:
		try:
			link = 'https://ru.glosbe.com/en/ru/{}'.format(word)
			word = translated_word(link)
			new_words.append(word)
			list_of_pk.append(pk)
			pk += 1
		except IndexError:
			pk += 1
			print('Error')

	return new_words


def write_to_excel_eng(eng_words):
	k = pk(english_words)

	for word in eng_words:
		sheet.cell(row=k, column=1).value = word
		k += 1

	wb.save("Eng.xlsx")


def write_to_excel_rus(rus_words):
	k = 0

	for several_words in rus_words:
		translate = ''
		index = len(several_words) - 1
		for word in several_words:
			if several_words[index] == word:
				translate += str(word)
			else:
				translate += str(word) + ', '

		sheet.cell(row=list_of_pk[k], column=3).value = translate
		k += 1

	wb.save("Eng.xlsx")


english_words = check_this_fucking_excel(english_words)  # Проверяю является ли пустая строка пустой -____-

russian_words = check_this_fucking_excel(russian_words)  # Проверяю является ли пустая строка пустой -____-

english_words = make_from_cell_list(english_words)  # Делаем из объектов excel список слов

english_words = filter_for_dot(english_words)  # Убираем символы, чтобы не мешали сравнению 

words = list(filter(filter_of_words, words))  # Проверка на реальные слова

words = filter_for_dot(words)  # Убираем символы, чтобы не мешали сравнению

words = cheak_repeat(words)  # Проверка на повторение спарсиных слов

words = cheak_repeat_in_excel(words, english_words)  # Проверка на повторение спарсиных слов и слов из excel

words = ask(words)  # Проверка на знание слов

words = words_from_longman(words)  # Поиск первой формы слова 

words = cheak_repeat(words)

words = cheak_repeat_in_excel(words, english_words)

translated_words = translated_words(words, pk(english_words))  # Перевод слов на русский

write_to_excel_eng(words)
write_to_excel_rus(translated_words)


print('Операция выполнена успешно! Иди проверь свой excel файлик)')

time.sleep(600)


